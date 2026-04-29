import csv
import io
import uuid
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.api.deps import RequestAuth, get_request_auth, require_module_access
from app.db.session import get_db
from app.models.akt_zatarki import AktZatarki
from app.models.applications import Application
from app.models.counterparty import Counterparty
from app.models.certificates_registry import CertificateRegistryRow
from app.models.registry_ui_settings import RegistryUiSetting
from app.schemas.certificates_registry import (
    CertificateRegistryColumnWidthsIn,
    CertificateRegistryColumnWidthsOut,
    CertificateRegistryImportOut,
    CertificateRegistryListOut,
    CertificateRegistryPatchIn,
    CertificateRegistryRowOut,
)

router = APIRouter(prefix="/certificates-registry", tags=["certificates-registry"])

MODULE_CODE = "registry_certificates"
WIDTHS_SETTING_KEY = "column_widths"


def _norm_header(value: str) -> str:
    return " ".join((value or "").replace("\n", " ").strip().upper().split())


HEADER_MAP = {
    "№": "registry_number",
    "BL": "bl_number",
    "BL ДАТА": "bl_date",
    "ВЕС": "weight_tons",
    "ФСС": "fss_number",
    "ДАТА ВЫДАЧИ ФСС": "fss_issue_date",
    "ФУМ": "fum",
    "СЕРТ КАЧЕСТВА": "quality_certificate",
    "ПИ": "pi",
    "ЗДОРОВЬЕ": "health",
    "ЗАКЛЮЧЕНИЕ": "conclusion",
    "РАДИО": "radio",
    "НОН ГМО": "non_gmo",
    "СОО": "soo",
    "WOOD": "wood",
}

REQUIRED_HEADERS = {"№"}


def _parse_date(value: str | None) -> date | None:
    if value is None:
        return None
    txt = str(value).strip()
    if not txt:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(txt, fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(value: str | None) -> float | None:
    if value is None:
        return None
    txt = str(value).strip().replace(",", ".")
    if not txt:
        return None
    try:
        return float(txt)
    except ValueError:
        return None


def _resolve_application_uuid(db: Session, registry_number: str) -> uuid.UUID | None:
    app = db.scalars(
        select(Application).where(Application.application_number == registry_number).limit(1)
    ).first()
    if app is not None:
        return app.uuid
    act = db.scalars(select(AktZatarki).where(AktZatarki.act_number == registry_number).limit(1)).first()
    if act is None:
        return None
    if act.application_uuid is not None:
        return act.application_uuid
    app_by_act = db.scalars(
        select(Application).where(Application.stuffing_act_uuid == act.uuid).limit(1)
    ).first()
    return app_by_act.uuid if app_by_act is not None else None


def _is_admin(request_auth: RequestAuth) -> bool:
    return request_auth.account is not None and request_auth.account.role_code == "admin"


def _get_default_counterparty_uuid(db: Session) -> uuid.UUID | None:
    row = db.scalars(
        select(Counterparty).where(Counterparty.is_active.is_(True)).order_by(Counterparty.created_at).limit(1)
    ).first()
    return row.uuid if row is not None else None


def _upsert_rows(db: Session, rows: list[dict[str, str]]) -> CertificateRegistryImportOut:
    processed = 0
    created = 0
    updated = 0
    linked = 0
    errors: list[str] = []
    existing_rows = list(db.scalars(select(CertificateRegistryRow)).all())
    by_number = {x.registry_number: x for x in existing_rows}
    for idx, src in enumerate(rows, start=2):
        registry_number = (src.get("registry_number") or "").strip()
        if not registry_number:
            errors.append(f"row {idx}: empty №")
            continue
        processed += 1
        row = by_number.get(registry_number)
        if row is None:
            row = CertificateRegistryRow(registry_number=registry_number)
            db.add(row)
            by_number[registry_number] = row
            created += 1
        else:
            updated += 1

        row.bl_number = (src.get("bl_number") or "").strip() or None
        row.bl_date = _parse_date(src.get("bl_date"))
        row.weight_tons = _parse_float(src.get("weight_tons"))
        row.fss_number = (src.get("fss_number") or "").strip() or None
        row.fss_issue_date = _parse_date(src.get("fss_issue_date"))
        row.fum = (src.get("fum") or "").strip() or None
        row.quality_certificate = (src.get("quality_certificate") or "").strip() or None
        row.pi = (src.get("pi") or "").strip() or None
        row.health = (src.get("health") or "").strip() or None
        row.conclusion = (src.get("conclusion") or "").strip() or None
        row.radio = (src.get("radio") or "").strip() or None
        row.non_gmo = (src.get("non_gmo") or "").strip() or None
        row.soo = (src.get("soo") or "").strip() or None
        row.wood = (src.get("wood") or "").strip() or None
        row.updated_at = datetime.utcnow()

        linked_uuid = _resolve_application_uuid(db, registry_number)
        if linked_uuid is not None:
            row.application_uuid = linked_uuid
            linked += 1
    db.commit()
    return CertificateRegistryImportOut(
        processed=processed, created=created, updated=updated, linked=linked, errors=errors
    )


@router.get("", response_model=CertificateRegistryListOut)
def list_registry_rows(
    db: Session = Depends(get_db),
    request_auth: RequestAuth = Depends(get_request_auth),
    q: str | None = Query(default=None),
) -> CertificateRegistryListOut:
    require_module_access(
        db=db,
        account=request_auth.account,
        roles=request_auth.roles,
        module_code=MODULE_CODE,
        need_write=False,
    )
    stmt = select(CertificateRegistryRow)
    if q and q.strip():
        token = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                CertificateRegistryRow.registry_number.ilike(token),
                CertificateRegistryRow.bl_number.ilike(token),
                CertificateRegistryRow.fss_number.ilike(token),
                CertificateRegistryRow.fum.ilike(token),
                CertificateRegistryRow.quality_certificate.ilike(token),
                CertificateRegistryRow.pi.ilike(token),
                CertificateRegistryRow.health.ilike(token),
                CertificateRegistryRow.conclusion.ilike(token),
                CertificateRegistryRow.radio.ilike(token),
                CertificateRegistryRow.non_gmo.ilike(token),
                CertificateRegistryRow.soo.ilike(token),
                CertificateRegistryRow.wood.ilike(token),
            )
        )
    stmt = stmt.order_by(CertificateRegistryRow.registry_number)
    items = list(db.scalars(stmt).all())
    return CertificateRegistryListOut(items=items, total=len(items))


@router.get("/column-widths", response_model=CertificateRegistryColumnWidthsOut)
def get_column_widths(
    db: Session = Depends(get_db),
    request_auth: RequestAuth = Depends(get_request_auth),
) -> CertificateRegistryColumnWidthsOut:
    require_module_access(
        db=db,
        account=request_auth.account,
        roles=request_auth.roles,
        module_code=MODULE_CODE,
        need_write=False,
    )
    row = db.get(RegistryUiSetting, WIDTHS_SETTING_KEY)
    widths = row.setting_value if row is not None and isinstance(row.setting_value, dict) else {}
    out: dict[str, int] = {}
    for key, value in widths.items():
        try:
            out[str(key)] = int(value)
        except (TypeError, ValueError):
            continue
    return CertificateRegistryColumnWidthsOut(widths=out)


@router.put("/column-widths", response_model=CertificateRegistryColumnWidthsOut)
def put_column_widths(
    payload: CertificateRegistryColumnWidthsIn,
    db: Session = Depends(get_db),
    request_auth: RequestAuth = Depends(get_request_auth),
) -> CertificateRegistryColumnWidthsOut:
    require_module_access(
        db=db,
        account=request_auth.account,
        roles=request_auth.roles,
        module_code=MODULE_CODE,
        need_write=True,
    )
    if not _is_admin(request_auth):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="admin_only")
    normalized: dict[str, int] = {}
    for key, value in payload.widths.items():
        width = max(80, min(1200, int(value)))
        normalized[str(key)] = width
    row = db.get(RegistryUiSetting, WIDTHS_SETTING_KEY)
    if row is None:
        row = RegistryUiSetting(setting_key=WIDTHS_SETTING_KEY, setting_value=normalized)
        db.add(row)
    else:
        row.setting_value = normalized
    db.commit()
    return CertificateRegistryColumnWidthsOut(widths=normalized)


@router.patch("/{row_uuid}", response_model=CertificateRegistryRowOut)
def patch_registry_row(
    row_uuid: uuid.UUID,
    payload: CertificateRegistryPatchIn,
    db: Session = Depends(get_db),
    request_auth: RequestAuth = Depends(get_request_auth),
) -> CertificateRegistryRowOut:
    require_module_access(
        db=db,
        account=request_auth.account,
        roles=request_auth.roles,
        module_code=MODULE_CODE,
        need_write=True,
    )
    row = db.get(CertificateRegistryRow, row_uuid)
    if row is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="registry_row_not_found")
    data = payload.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="empty_patch")
    for key, value in data.items():
        setattr(row, key, value)
    row.updated_at = datetime.utcnow()
    if row.registry_number:
        row.application_uuid = _resolve_application_uuid(db, row.registry_number)
    db.commit()
    db.refresh(row)
    return row


@router.post("/import", response_model=CertificateRegistryImportOut)
def import_registry(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    request_auth: RequestAuth = Depends(get_request_auth),
) -> CertificateRegistryImportOut:
    require_module_access(
        db=db,
        account=request_auth.account,
        roles=request_auth.roles,
        module_code=MODULE_CODE,
        need_write=True,
    )
    already = int(db.scalar(select(func.count()).select_from(CertificateRegistryRow)) or 0)
    if already > 0:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="import_already_completed",
        )

    content = file.file.read()
    filename = (file.filename or "").lower()
    parsed_rows: list[dict[str, str]] = []
    headers: list[str] = []
    if filename.endswith(".csv"):
        txt = content.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(io.StringIO(txt), delimiter=";")
        matrix = list(reader)
        if matrix and len(matrix[0]) < 3:
            reader = csv.reader(io.StringIO(txt), delimiter=",")
            matrix = list(reader)
        if not matrix:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="empty_file")
        headers = [_norm_header(h) for h in matrix[0]]
        for row in matrix[1:]:
            parsed_rows.append({headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))})
    elif filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="empty_file")
        headers = [_norm_header(str(x or "")) for x in values[0]]
        for r in values[1:]:
            parsed_rows.append({headers[i]: ("" if r[i] is None else str(r[i])) for i in range(len(headers))})
    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="unsupported_file_type")

    missing = [x for x in REQUIRED_HEADERS if x not in headers]
    if missing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"missing_headers:{','.join(missing)}")

    normalized_rows: list[dict[str, str]] = []
    for src in parsed_rows:
        out: dict[str, str] = {}
        for raw_key, value in src.items():
            mapped = HEADER_MAP.get(raw_key)
            if mapped:
                out[mapped] = value
        if out:
            normalized_rows.append(out)
    return _upsert_rows(db, normalized_rows)


@router.post("/backfill")
def backfill_from_registry(
    db: Session = Depends(get_db),
    request_auth: RequestAuth = Depends(get_request_auth),
) -> dict:
    require_module_access(
        db=db,
        account=request_auth.account,
        roles=request_auth.roles,
        module_code=MODULE_CODE,
        need_write=True,
    )
    default_counterparty_uuid = _get_default_counterparty_uuid(db)
    if default_counterparty_uuid is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="no_active_counterparty_for_act_creation")

    rows = list(db.scalars(select(CertificateRegistryRow).order_by(CertificateRegistryRow.registry_number)).all())
    created_apps = 0
    created_acts = 0
    updated_weights = 0
    linked = 0
    errors: list[str] = []

    for row in rows:
        number = (row.registry_number or "").strip()
        if not number:
            errors.append(f"row:{row.uuid}:empty_number")
            continue
        app = db.scalars(
            select(Application).where(Application.application_number == number).limit(1)
        ).first()
        if app is None:
            app = Application(
                source_type="manual",
                status_code="DRAFT",
                application_number=number,
                weight_tons=row.weight_tons,
            )
            db.add(app)
            db.flush()
            created_apps += 1
        else:
            if row.weight_tons is not None and app.weight_tons != row.weight_tons:
                app.weight_tons = row.weight_tons
                updated_weights += 1

        act = db.scalars(select(AktZatarki).where(AktZatarki.act_number == number).limit(1)).first()
        if act is None:
            act = AktZatarki(
                act_number=number,
                act_date=date.today(),
                product_name_ru=None,
                counterparty_uuid=default_counterparty_uuid,
                source="manual",
                status="active",
                application_uuid=app.uuid,
            )
            db.add(act)
            db.flush()
            created_acts += 1
        else:
            act.application_uuid = app.uuid

        app.stuffing_act_uuid = act.uuid
        row.application_uuid = app.uuid
        row.updated_at = datetime.utcnow()
        linked += 1

    db.commit()
    return {
        "processed": len(rows),
        "created_applications": created_apps,
        "created_acts": created_acts,
        "updated_weights": updated_weights,
        "linked_rows": linked,
        "errors": errors,
    }


@router.post("/relink")
def relink_registry_rows(
    db: Session = Depends(get_db),
    request_auth: RequestAuth = Depends(get_request_auth),
) -> dict:
    require_module_access(
        db=db,
        account=request_auth.account,
        roles=request_auth.roles,
        module_code=MODULE_CODE,
        need_write=True,
    )
    rows = list(db.scalars(select(CertificateRegistryRow)).all())
    linked = 0
    unlinked = 0
    for row in rows:
        linked_uuid = _resolve_application_uuid(db, row.registry_number)
        row.application_uuid = linked_uuid
        row.updated_at = datetime.utcnow()
        if linked_uuid is None:
            unlinked += 1
        else:
            linked += 1
    db.commit()
    return {"processed": len(rows), "linked": linked, "unlinked": unlinked}
